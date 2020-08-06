import openpyxl as xl
from openpyxl.chart import BarChart, Reference

filename = input(Enter file name: )
coloum_number = input(Enter coloum: )
multiplyer = input(Enter the multiplyer: )

def price_changer(filename,coloum_number,multiplyer):
    wd= xl.load_workbook(filename)
    sheet = wd['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, coloum_number)
        correct_price = cell.value * multiplyer
        cell.value = correct_price

    value = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'a5')

    wd.save(filename)

price_changer(filename,coloum_number,multiplyer)
